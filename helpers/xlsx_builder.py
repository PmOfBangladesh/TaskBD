# ============================================================
#  helpers/xlsx_builder.py  —  XLSX generation
# ============================================================
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_THIN   = Border(left=Side(style="thin"), right=Side(style="thin"),
                 top=Side(style="thin"),  bottom=Side(style="thin"))
_CENTER = Alignment(horizontal="center", vertical="center")


def build_report_xlsx(rows: list[dict], title: str = "Report",
                      headers: list[str] = None) -> io.BytesIO:
    wb  = Workbook()
    ws  = wb.active
    ws.title = "Data"

    if not rows:
        ws["A1"] = "No data."
        buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

    if not headers:
        headers = list(rows[0].keys())

    col_n = len(headers)

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(col_n)}1")
    ws["A1"] = title
    ws["A1"].font      = Font(bold=True, color="FFFFFF", size=13, name="Arial")
    ws["A1"].fill      = PatternFill("solid", fgColor="0D3362")
    ws["A1"].alignment = _CENTER
    ws.row_dimensions[1].height = 28

    # Timestamp row
    ws.merge_cells(f"A2:{get_column_letter(col_n)}2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font      = Font(italic=True, color="595959", size=10, name="Arial")
    ws["A2"].alignment = _CENTER

    # Header row
    hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.border    = _THIN
        c.alignment = _CENTER
    ws.row_dimensions[3].height = 20

    # Data rows
    alt = PatternFill("solid", fgColor="D9E1F2")
    wht = PatternFill("solid", fgColor="FFFFFF")
    for ri, row in enumerate(rows, 4):
        fill = alt if ri % 2 == 0 else wht
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=ri, column=ci, value=row.get(h, ""))
            c.font      = Font(name="Arial", size=10)
            c.fill      = fill
            c.border    = _THIN
            c.alignment = _CENTER

    # Column widths
    for ci, h in enumerate(headers, 1):
        w = max(len(str(h)), max((len(str(r.get(h, ""))) for r in rows), default=0))
        ws.column_dimensions[get_column_letter(ci)].width = min(w + 4, 40)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf


def build_plain_xlsx(rows: list[dict], headers: list[str]) -> io.BytesIO:
    wb = Workbook(); ws = wb.active; ws.title = "Accounts"
    for i, row in enumerate(rows, 1):
        for ci, h in enumerate(headers, 1):
            ws.cell(row=i, column=ci, value=row.get(h, "")).font = Font(name="Calibri", size=10)
    for ci, h in enumerate(headers, 1):
        vals = [str(r.get(h, "")) for r in rows]
        ws.column_dimensions[get_column_letter(ci)].width = min(max((len(v) for v in vals), default=8) + 2, 45)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf
